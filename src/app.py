import os
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, Equipment, Employee, Document, Department, EquipmentImei, EquipmentPhone, Movement, AssignmentHistory
from dotenv import load_dotenv
from flask import send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import load_workbook  
from datetime import datetime, timezone
import openpyxl
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from sqlalchemy import func

load_dotenv()

app = Flask(__name__)

# --- Configurações de Pastas ---
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')

db.init_app(app)

with app.app_context():
    db.create_all()
    print("Banco de dados verificado/criado com sucesso!")

# --- Configuração do Flask-Login ---
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Rotas de Autenticação ---

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role') # 'RH' ou 'TI'

        user_exists = User.query.filter_by(username=username).first()
        if user_exists:
            flash('Usuário já existe!')
            return redirect(url_for('cadastro'))

        # Hash da senha por segurança
        hashed_pw = generate_password_hash(password, method='pbkdf2:sha256')
        
        new_user = User(username=username, password=hashed_pw, role=role)
        db.session.add(new_user)
        db.session.commit()
        
        flash('Usuário criado com sucesso!')
        return redirect(url_for('login'))

    return render_template('cadastro.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('dashboard')) # Rota principal que você criará
        else:
            flash('Login inválido. Verifique suas credenciais.')
            
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    try:
        # 1. Contadores de Recursos Humanos
        ativos = Employee.query.filter_by(status='Active').count()
        desligados = Employee.query.filter_by(status='Terminated').count()
        termos = Document.query.count() if 'Document' in globals() else 0

        # 2. Contadores do Inventário de TI
        total_equip = Equipment.query.count()
        disponiveis = Equipment.query.filter_by(inventory_status='Available').count()
        em_uso = Equipment.query.filter_by(inventory_status='Assigned').count()
        em_revisao = Equipment.query.filter_by(inventory_status='Under Review').count()

        # 3. Query Corrigida com OUTERJOIN (Evita quebrar se não houver funcionários ou departamentos cadastrados)
        dados_deptos = db.session.query(
            Department.name, 
            func.count(Employee.id)
        ).outerjoin(Employee, Employee.department_id == Department.id)\
         .filter((Employee.status == 'Active') | (Employee.id == None))\
         .group_by(Department.name)\
         .order_by(func.count(Employee.id).desc())\
         .all()
         
         # Limpa registros nulos gerados pelo outerjoin caso o banco esteja 100% zerado
        dados_deptos = [(depto, qtd) for depto, qtd in dados_deptos if depto is not None]

    except Exception as e:
        print(f"Erro interno no processamento do Dashboard: {e}")
        ativos = desligados = termos = total_equip = disponiveis = em_uso = em_revisao = 0
        dados_deptos = []

    return render_template(
        'dashboard.html',
        ativos=ativos,
        desligados=desligados,
        termos=termos,
        total_equip=total_equip,
        disponiveis=disponiveis,
        em_uso=em_uso,
        em_revisao=em_revisao,
        dados_deptos=dados_deptos
    )

@app.route('/estoque')
@login_required
def estoque():
    # Busca todos os itens do banco
    itens = Equipment.query.all()
    return render_template('estoque.html', itens=itens)

@app.route('/estoque/novo', methods=['POST'])
@login_required
def novo_item():
    item_type = request.form.get('item_type')
    model = request.form.get('model')
    serial_number = request.form.get('serial_number')
    
    # Captura os dados do formulário
    phone_number = request.form.get('phone_number') if item_type == 'Celular' else None
    imei = request.form.get('imei') if item_type == 'Celular' else None

    # Verifica se o serial já existe
    exists = Equipment.query.filter_by(serial_number=serial_number).first()
    if exists:
        flash('Erro: Já existe um equipamento com este Número de Série!', 'danger')
    else:
        try:
            # 1. Cria e salva o Equipamento Principal primeiro
            novo = Equipment(
                item_type=item_type,
                model=model,
                serial_number=serial_number,
                inventory_status='Available'
            )
            db.session.add(novo)
            
            # Dá o flush para gerar o ID do equipamento antes de commitar tudo
            db.session.flush() 

            # 2. Se for Celular e os campos foram preenchidos, vincula os registros extras
            if item_type == 'Celular':
                if phone_number and phone_number.strip():
                    novo_telefone = EquipmentPhone(
                        phone_number=phone_number.strip(),
                        equipment_id=novo.id  # Usa o ID recém-gerado
                    )
                    db.session.add(novo_telefone)

                if imei and imei.strip():
                    novo_imei = EquipmentImei(
                        imei_value=imei.strip(),  # Ajustado para o nome do campo na sua Model (imei_value)
                        equipment_id=novo.id   # Usa o ID recém-gerado
                    )
                    db.session.add(novo_imei)

            # 3. Commita todas as transações juntas com segurança
            db.session.commit()
            flash('Item adicionado ao estoque!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar no banco de dados: {str(e)}', 'danger')
    
    return redirect(url_for('estoque'))

# Rota para Liberar Equipamento em Revisão / Desvincular
@app.route('/estoque/liberar/<int:id>', methods=['POST'])
@login_required
def liberar_equipamento(id):
    equip = Equipment.query.get_or_404(id)
    descricao_devolucao = request.form.get('descricao_devolucao', '').strip()
    
    try:
        dono_historico = "Disponível (Sem Dono)"
        acao_movimento = "Liberado do estoque"
        
        if equip.owner:
            dono_historico = f"Ex-colaborador: {equip.owner.name} (CPF: {equip.owner.cpf})"
            acao_movimento = "Liberado após desligamento"
        elif equip.department_owner:
            dono_historico = f"Uso Comum anterior: {equip.department_owner.name}"
            acao_movimento = "Liberado de uso coletivo do setor"

        if descricao_devolucao:
            dono_historico += f" | Obs técnica: {descricao_devolucao}"

        # -------------------------------------------------------------
        # REGRA ATUALIZADA: Finaliza o histórico de posse ativo
        # -------------------------------------------------------------
        historico_ativo = AssignmentHistory.query.filter_by(
            equipment_id=equip.id, 
            returned_at=None
        ).first()
        
        if historico_ativo:
            historico_ativo.returned_at = datetime.now(timezone.utc)
            historico_ativo.status = 'Returned'
            historico_ativo.notes = descricao_devolucao
        # -------------------------------------------------------------

        # Alimenta a tabela geral de auditoria rápida/Movimentos que você já possui
        historico_mov = Movement(
            equipment_id=equip.id,
            equipment_model=equip.model,
            serial_number=equip.serial_number,
            previous_user=dono_historico,
            action=acao_movimento
        )
        db.session.add(historico_mov)

        # Limpa o equipamento para o estoque
        equip.inventory_status = 'Available'
        equip.employee_id = None
        equip.department_id = None
        equip.assignment_date = None
        
        db.session.commit()
        flash(f'O equipamento "{equip.model}" foi liberado com sucesso e o histórico de posse foi preservado.')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao liberar equipamento: {str(e)}')
        
    return redirect(url_for('estoque'))

@app.route('/funcionarios/importar', methods=['POST'])
@login_required
def importar_excel_funcionarios():
    if 'file' not in request.files:
        flash('Nenhum arquivo enviado.')
        return redirect(url_for('funcionarios'))
        
    file = request.files['file']
    if file.filename == '':
        flash('Nenhum arquivo selecionado.')
        return redirect(url_for('funcionarios'))

    # Altera a validação para aceitar arquivos nativos do Excel
    if file and (file.filename.lower().endswith('.xlsx') or file.filename.lower().endswith('.xls')):
        try:
            # Carrega a planilha diretamente da memória
            wb = openpyxl.load_workbook(file.stream, data_only=True)
            sheet = wb.active  # Pega a primeira aba da planilha

            # 1. Capturar os cabeçalhos da primeira linha (tira espaços e joga para minúsculo)
            headers = [str(cell.value).strip().lower() if cell.value is not None else None for cell in sheet[1]]
            
            sucesso = 0

            # 2. Iterar pelas linhas de dados (começando da linha 2)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Se a linha estiver completamente vazia, pula
                if not any(row):
                    continue

                # Monta um dicionário seguro mapeando cabeçalho -> valor da célula
                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx] is not None:
                        row_data[headers[col_idx]] = str(value).strip() if value is not None else ''

                # Coleta os dados usando os nomes exatos das colunas em inglês do seu modelo
                name = row_data.get('name', '')
                cpf = row_data.get('cpf', '')
                dept_name = row_data.get('department', '')
                status = row_data.get('status', 'Active')
                
                # Dados do equipamento
                item_type = row_data.get('item_type', '')
                model = row_data.get('model', '')
                serial_number = row_data.get('serial_number', '')
                phone_number = row_data.get('phone_number', '')
                imei_value = row_data.get('imei', '')

                # Se a linha não tiver nome de funcionário ou setor, ignora
                if not name:
                    continue

                # 1. Trata a existência do Departamento no banco
                depto = None
                if dept_name:
                    depto = Department.query.filter_by(name=dept_name).first()
                    if not depto:
                        depto = Department(name=dept_name)
                        db.session.add(depto)
                        db.session.flush()

                funcionario_id = None
                departamento_id = depto.id if depto else None

                # REGRA DE OURO: Se Name e Department forem iguais E não houver CPF -> USO COMUM
                eh_uso_comum = (name.lower() == dept_name.lower()) and (not cpf)

                if not eh_uso_comum:
                    # Busca ou cria o funcionário se não for uso comum
                    func = Employee.query.filter_by(cpf=cpf).first() if cpf else None
                    
                    if not func:
                        func = Employee(
                            name=name,
                            cpf=cpf if cpf else f"SEM_CPF_{sucesso}_{datetime.now().microsecond}",
                            status=status if status else 'Active',
                            department_id=departamento_id
                        )
                        db.session.add(func)
                        db.session.flush()
                    
                    funcionario_id = func.id
                    departamento_id = None  # Equipamento fica atrelado à pessoa física

                # 2. Processa o Equipamento/Ativo
                if model:
                    equip = Equipment.query.filter_by(serial_number=serial_number).first() if serial_number else None
                    
                    if not equip:
                        status_inventario = 'Assigned'
                        if status == 'Terminated':
                            status_inventario = 'Under Review'

                        equip = Equipment(
                            model=model,
                            serial_number=serial_number if serial_number else None,
                            item_type=item_type if item_type else 'Não Definido',
                            inventory_status=status_inventario,
                            employee_id=funcionario_id,
                            department_id=departamento_id,
                            assignment_date=datetime.now(timezone.utc)
                        )
                        db.session.add(equip)
                        db.session.flush()

                        historico_posse = AssignmentHistory(
                            equipment_id=equip.id,
                            employee_id=funcionario_id, # Se houver
                            department_id=departamento_id, # Se for uso comum do setor
                            assigned_at=datetime.now(timezone.utc),
                            status='Active'
                        )
                        db.session.add(historico_posse)

                        # Se for celular, alimenta as tabelas complementares
                        if item_type and 'celular' in item_type.lower():
                            if phone_number:
                                novo_tel = EquipmentPhone(phone_number=phone_number, equipment_id=equip.id)
                                db.session.add(novo_tel)
                            if imei_value:
                                novo_imei = EquipmentImei(imei_value=imei_value, equipment_id=equip.id)
                                db.session.add(novo_imei)

                sucesso += 1

            db.session.commit()
            flash(f'Importação concluída com sucesso! {sucesso} registros processados.')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro crítico durante a importação da planilha: {str(e)}')
            print(f"Erro detalhado na importação XLSX: {e}")
            
    else:
        flash('Formato de arquivo inválido. Selecione uma planilha do Excel (.xlsx ou .xls)')

    return redirect(url_for('funcionarios'))

@app.route('/estoque/excluir/<int:id>')
@login_required
def excluir_item(id):
    item = Equipment.query.get_or_404(id)
    # Impede exclusão se estiver com um funcionário (opcional, por segurança)
    if item.employee_id:
        flash('Não é possível excluir um item que está em posse de um funcionário!')
    else:
        db.session.delete(item)
        db.session.commit()
        flash('Item removido com sucesso.')
    return redirect(url_for('estoque'))

@app.route('/departamentos/novo', methods=['POST'])
@login_required
def novo_departamento():
    name = request.form.get('name', '').strip()
    
    if not name:
        flash('Erro: O nome do departamento não pode ser vazio.')
        return redirect(url_for('funcionarios'))
        
    # CORREÇÃO AQUI: Mudado de ilocike para ilike
    depto_existente = Department.query.filter(Department.name.ilike(name)).first()
    if depto_existente:
        flash(f'O departamento "{name}" já está cadastrado.')
        return redirect(url_for('funcionarios'))
        
    try:
        novo_depto = Department(name=name)
        db.session.add(novo_depto)
        db.session.commit()
        flash(f'Departamento "{name}" cadastrado com sucesso!')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao salvar departamento: {str(e)}')
        
    return redirect(url_for('funcionarios'))

@app.route('/novo_equipamento_comum', methods=['POST'])
@login_required
def novo_equipamento_comum():
    # Pega o id do equipamento vindo do select de estoque (Função 1)
    estoque_equipment_id = request.form.get('estoque_equipment_id')
    department_id = request.form.get('department_id')

    if not department_id:
        flash('Por favor, selecione um departamento responsável.', 'danger')
        return redirect(url_for('funcionarios')) # Ou a rota onde está o seu modal

    # CAMINHO A: O usuário escolheu um equipamento já existente no estoque
    if estoque_equipment_id:
        equipment = Equipment.query.get(estoque_equipment_id)
        
        if equipment:
            # Apenas atrela ao setor selecionado
            equipment.department_id = department_id
            equipment.inventory_status = "Assigned"
            
            # Se você tiver controle de status para saber que é de uso comum, atualize aqui.
            # Exemplo: equipment.is_common_use = True
            
            # (Opcional) Registrar no histórico de movimentações que ele foi para uso comum
            # nova_movimentacao = Movimentacao(equipment_id=equipment.id, ... )
            
            db.session.commit()
            flash(f'Equipamento {equipment.model} atrelado ao setor com sucesso!', 'success')
        else:
            flash('Equipamento selecionado do estoque não foi encontrado.', 'danger')
            
        return redirect(url_for('funcionarios'))

    # CAMINHO B: Cadastro Manual (Caso o usuário não tenha selecionado nada do estoque)
    item_type = request.form.get('item_type')
    model = request.form.get('model')
    serial_number = request.form.get('serial_number', '').strip() or None

    # Validação de S/N duplicado apenas para novos cadastros
    if serial_number:
        existing_equip = Equipment.query.filter_by(serial_number=serial_number).first()
        if existing_equip:
            flash(f'Erro: Já existe um equipamento cadastrado com o S/N: {serial_number}!', 'danger')
            return redirect(url_for('funcionarios'))

    # Criação do novo equipamento do zero
    novo_equipamento = Equipment(
        item_type=item_type,
        model=model,
        serial_number=serial_number,
        department_id=department_id,
        # Se for celular manual, pega os dados adicionais (Função 2)
        phone_number=request.form.get('phone_number') if item_type.lower() == 'celular' else None,
        imei=request.form.get('imei') if item_type.lower() == 'celular' else None
    )

    try:
        db.session.add(novo_equipamento)
        db.session.commit()
        flash('Novo equipamento comum cadastrado e atrelado ao setor com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao salvar no banco de dados: {str(e)}', 'danger')

    return redirect(url_for('funcionarios'))

@app.route('/estoque/editar/<int:id>', methods=['POST'])
@login_required
def editar_equipamento(id):
    equip = Equipment.query.get_or_404(id)
    
    item_type = request.form.get('item_type', '').strip()
    model = request.form.get('model', '').strip()
    serial_number = request.form.get('serial_number', '').strip() or None
    inventory_status = request.form.get('inventory_status', equip.inventory_status)
    
    phone_number_raw = request.form.get('phone_number', '').strip()
    imei_raw = request.form.get('imei', '').strip()

    if not item_type or not model:
        flash('Erro: Tipo e Modelo não podem ficar vazios.')
        return redirect(url_for('estoque'))

    # Valida duplicidade de S/N se ele foi alterado
    if serial_number and serial_number != equip.serial_number:
        existente = Equipment.query.filter_by(serial_number=serial_number).first()
        if existente:
            flash(f'Erro: O Número de Série {serial_number} já pertence a outro ativo.')
            return redirect(url_for('estoque'))

    try:
        # Atualiza dados principais
        equip.item_type = item_type
        equip.model = model
        equip.serial_number = serial_number
        equip.inventory_status = inventory_status
        
        # Se mudar para disponível, rompe vínculos antigos por segurança
        if inventory_status == 'Available':
            equip.employee_id = None
            equip.department_id = None

        # Atualiza múltiplos chips se for um celular
        if phone_number_raw:
            EquipmentPhone.query.filter_by(equipment_id=equip.id).delete()
            telefones = [t.strip() for t in phone_number_raw.split('/') if t.strip()]
            for tel in telefones:
                db.session.add(EquipmentPhone(phone_number=tel, equipment_id=equip.id))

        # Atualiza múltiplos IMEIs se informado
        if imei_raw:
            EquipmentImei.query.filter_by(equipment_id=equip.id).delete()
            imeis = [i.strip() for i in imei_raw.split('/') if i.strip()]
            for imei in imeis:
                db.session.add(EquipmentImei(imei_value=imei, equipment_id=equip.id))

        db.session.commit()
        flash(f'Equipamento {equip.model} atualizado com sucesso!')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar equipamento: {str(e)}')

    return redirect(url_for('estoque'))

@app.route('/estoque/excluir/<int:id>', methods=['POST', 'GET'])
@login_required
def excluir_equipamento(id):
    equip = Equipment.query.get_or_404(id)
    
    try:
        modelo_removido = equip.model
        db.session.delete(equip)
        db.session.commit()
        flash(f'Equipamento {modelo_removido} removido do inventário permanente.')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir equipamento: {str(e)}')
        
    return redirect(url_for('estoque'))

@app.route('/funcionarios')
@login_required
def funcionarios():
    # Pega o filtro da URL (padrão é 'Active' se não for especificado)
    status_filtro = request.args.get('status', 'Active')

    lista_departamentos = Department.query.all()
    
    # Base da Query
    query = Employee.query.filter_by(status=status_filtro)
    
    # Aplica a ordenação baseada no filtro selecionado
    if status_filtro == 'Terminated':
        # Se for desligado, ordena pela data de desligamento decrescente
        lista = query.order_by(Employee.termination_date.desc()).all()
    else:
        # Se for ativo (ou qualquer outro), ordena pela data de contratação decrescente
        lista = query.order_by(Employee.hiring_date.desc()).all()
        
    equipamentos_livres = Equipment.query.filter_by(inventory_status='Available').all()
    
    # Passamos o status_filtro para o HTML saber qual filtro está ativo na tela
    return render_template('funcionarios.html', 
                           funcionarios=lista, 
                           equipamentos_livres=equipamentos_livres, 
                           current_filter=status_filtro,
                           departamentos=lista_departamentos)

# Rota para Criar Funcionário
@app.route('/funcionarios/novo', methods=['POST'])
@login_required
def novo_funcionario():
    nome = request.form.get('name')
    cpf = request.form.get('cpf')
    email = request.form.get('email')
    depto = request.form.get('department')
    
    novo = Employee(name=nome, cpf=cpf, email=email, department=depto)
    db.session.add(novo)
    db.session.commit()
    flash('Funcionário cadastrado!')
    return redirect(url_for('funcionarios'))


@app.route('/estoque/desvincular/<int:id>', methods=['POST'])
@login_required
def desvincular_equipamento(id):
    equip = Equipment.query.get_or_404(id)
    origem = request.args.get('origem', 'estoque')
    
    try:
        # Identifica o dono atual para o histórico antes de apagar
        dono_anterior = "Disponível (Sem Dono)"
        if equip.owner:
            dono_anterior = f"Colaborador: {equip.owner.name} (CPF: {equip.owner.cpf})"
        elif equip.department_owner:
            dono_anterior = f"Uso Comum: {equip.department_owner.name}"

        # 1. Cria o registro de movimentação histórica
        historico = Movement(
            equipment_id=equip.id,
            equipment_model=equip.model,
            serial_number=equip.serial_number,
            previous_user=dono_anterior,
            action="Liberado para Estoque"
        )
        db.session.add(historico)

        # 2. SEPARAÇÃO MANUAL COMPLETA (Zera todas as chaves estrangeiras)
        equip.employee_id = None
        equip.department_id = None
        equip.inventory_status = 'Available'
        equip.assignment_date = None
        
        # Força o SQLAlchemy a registrar a alteração de nulo
        db.session.flush()
        db.session.commit()
        
        flash(f'O equipamento "{equip.model}" foi desvinculado e retornou ao estoque disponível.')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao desvincular equipamento: {str(e)}')
        print(f"Erro detalhado na desvinculação: {e}")
        
    return redirect(url_for(origem))

@app.route('/funcionarios/gerar_pdf/<int:id>')
@login_required
def gerar_pdf_funcionario(id):
    func = Employee.query.get_or_404(id)
    
    # Criar um buffer na memória para receber o PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54)
    story = []
    
    # Configuração de Estilos do Texto
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'],
        alignment=TA_CENTER, fontSize=14, leading=18, spaceAfter=30
    )
    
    body_style = ParagraphStyle(
        'BodyStyle', parent=styles['Normal'],
        alignment=TA_JUSTIFY, fontSize=11, leading=16, spaceAfter=20
    )
    
    meta_style = ParagraphStyle(
        'MetaStyle', parent=styles['Normal'],
        alignment=TA_LEFT, fontSize=11, leading=16, spaceAfter=30
    )

    # 1. LOGO CENTRALIZADA (Usando texto estilizado caso a imagem falhe no mini-servidor)
    logo_style = ParagraphStyle('LogoStyle', alignment=TA_CENTER, fontSize=24, leading=28, textColor=colors.HexColor('#1a365d'), spaceAfter=20)
    story.append(Paragraph("<b>BTL Cargo LTDA</b>", logo_style))
    story.append(Spacer(1, 15))
    
    # 2. TÍTULO DO TERMO
    titulo = "TERMO DE RESPONSABILIDADE DE EQUIPAMENTOS DE INFORMÁTICA"
    story.append(Paragraph(f"<b>{titulo}</b>", title_style))
    
    # 3. TEXTO DECLARATÓRIO
    texto_declaracao = (
        f"Eu, <b>{func.name}</b>, portador(a) do CPF nº <b>{func.cpf}</b>, "
        f"colaborador(a) da empresa BTL Cargo LTDA, declaro que recebi os equipamentos "
        f"abaixo relacionados, em perfeito estado de funcionamento:"
    )
    story.append(Paragraph(texto_declaracao, body_style))
    
    # 4. TABELA DE EQUIPAMENTOS
    dados_tabela = [["Equipamento / Modelo", "Número de Série (S/N)"]]
    for eq in func.equipments:
        dados_tabela.append([f"{eq.item_type} - {eq.model}", eq.serial_number])
        
    # Se o funcionário não tiver equipamentos atribuídos ainda
    if len(dados_tabela) == 1:
        dados_tabela.append(["Nenhum equipamento vinculado no sistema.", "---"])
        
    tabela = Table(dados_tabela, colWidths=[280, 220])
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(tabela)
    story.append(Spacer(1, 40))
    
    agora = datetime.now(timezone.utc)
    dia = agora.strftime('%d')
    ano = agora.strftime('%Y')
    
    # Mapeamento manual dos meses para garantir PT-BR independente do servidor
    meses_pt = {
        'January': 'de Janeiro', 'February': 'de Fevereiro', 'March': 'de Março',
        'April': 'de Abril', 'May': 'de Maio', 'June': 'de Junho',
        'July': 'de Julho', 'August': 'de Agosto', 'September': 'de Setembro',
        'October': 'de Outubro', 'November': 'de Novembro', 'December': 'de Dezembro'
    }
    
    mes_ingles = agora.strftime('%B')
    mes_portugues = meses_pt.get(mes_ingles, mes_ingles) # Caso falhe, mantém o original
    
    texto_data = f"Guarulhos, SP, {dia} {mes_portugues} de {ano}."
    story.append(Paragraph(texto_data, meta_style))
    story.append(Spacer(1, 40))
    
    # 6. ESPAÇO PARA ASSINATURAS
    dados_assinatura = [
        ["_______________________________________", "_______________________________________"],
        [func.name, current_user.username],
        ["Colaborador(a)", "Responsável pela Entrega"]
    ]
    tabela_assinatura = Table(dados_assinatura, colWidths=[250, 250])
    tabela_assinatura.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
    ]))
    story.append(tabela_assinatura)
    
    # Construir o PDF
    doc.build(story)
    buffer.seek(0)
    
    # Envia o arquivo de volta limpando o nome do arquivo gerado
    nome_arquivo_pdf = f"Termo_{func.name.replace(' ', '_')}.pdf"
    return send_file(buffer, as_attachment=False, mimetype='application/pdf', download_name=nome_arquivo_pdf)

# Rota para Editar/Vincular Termo e Equipamento
@app.route('/funcionarios/editar/<int:id>', methods=['POST'])
@login_required
def editar_funcionario(id):
    func = Employee.query.get_or_404(id)
    func.name = request.form.get('name')
    
    # Processar múltiplos uploads de termos
    if 'termo' in request.files:
        file = request.files['termo']
        if file.filename != '':
            desc_termo = request.form.get('desc_termo') or "Termo de Responsabilidade"
            
            # Nome único usando timestamp para evitar sobrescrever arquivos
            timestamp = int(datetime.now(timezone.utc).timestamp())
            filename = secure_filename(f"termo_{func.id}_{timestamp}_{file.filename}")
            
            # 1. DEFINE O CAMINHO DA SUBPASTA 'termos'
            pasta_termos = os.path.join(app.config['UPLOAD_FOLDER'], 'termos')
            
            # 2. GARANTE QUE A PASTA 'termos' EXISTE (Se não existir, o Python cria)
            os.makedirs(pasta_termos, exist_ok=True)
            
            # 3. JUNTA TUDO PARA GERAR O CAMINHO FÍSICO FINAL DO ARQUIVO
            filepath = os.path.join(pasta_termos, filename)
            file.save(filepath)
            
            # Salva o novo documento na tabela secundária
            novo_documento = Document(filename=filename, description=desc_termo, employee_id=func.id)
            db.session.add(novo_documento)

    # Vincular Equipamento
    eq_id = request.form.get('equipamento_id')
    if eq_id:
        equip = Equipment.query.get(eq_id)
        equip.employee_id = func.id
        equip.inventory_status = 'Assigned'
        equip.assignment_date = datetime.now(timezone.utc)

    db.session.commit()
    flash('Cadastro atualizado com sucesso!')
    return redirect(url_for('funcionarios'))

# Atualização da Rota de Desligamento
@app.route('/funcionarios/desligar/<int:id>', methods=['POST'])
@login_required
def desligar_funcionario(id):
    func = Employee.query.get_or_404(id)
    
    try:
        func.status = 'Terminated'
        func.termination_date = datetime.now(timezone.utc)
        
        # Caça todos os equipamentos que estavam com ele
        for equip in func.equipments:
            # 1. Registra no histórico que o item foi retido para revisão devido à demissão
            mov_retencao = Movement(
                equipment_id=equip.id,
                equipment_model=equip.model,
                serial_number=equip.serial_number,
                previous_user=f"Colaborador Desligado: {func.name} (CPF: {func.cpf})",
                action="Retido para Revisão (Desligamento)"
            )
            db.session.add(mov_retencao)
            
            # 2. Muda o status do item, mas NÃO limpa o employee_id ainda!
            equip.inventory_status = 'Under Review'
            
        db.session.commit()
        flash(f'Funcionário {func.name} desligado com sucesso. Equipamentos retidos para revisão.')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao desligar funcionário: {str(e)}')
        
    return redirect(url_for('funcionarios'))

@app.route('/movimentacoes')
@login_required
def listar_movimentacoes():
    # Busca todas as movimentações ordenando da mais recente para a mais antiga
    todas_movimentacoes = Movement.query.order_by(Movement.date.desc()).all()
    return render_template('movements.html', movements=todas_movimentacoes)

# Rota para baixar/ver o termo
@app.route('/uploads/termos/<filename>')
@login_required
def visualizar_termo(filename):
    # Constrói o caminho absoluto para a pasta 'uploads/termos'
    diretorio_termos = os.path.join(app.config['UPLOAD_FOLDER'], 'termos')
    
    # Envia o arquivo de forma segura para o navegador
    return send_from_directory(diretorio_termos, filename)

if __name__ == '__main__':
    # Rodando em 0.0.0.0 para ser acessível na rede local
    with app.app_context():
        db.create_all()
    app.run(host='0.0.0.0', port=5000, debug=True)